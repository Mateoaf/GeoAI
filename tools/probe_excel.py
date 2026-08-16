from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data" / "raw" / "oxrep-mines-3.0-20250408.xlsx"


def normalise_scalar(value):
    if pd.isna(value):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    return value


def main() -> None:
    workbook = load_workbook(INPUT, data_only=False, read_only=False)
    workbook_meta = []
    for sheet in workbook.worksheets:
        formula_count = 0
        nonempty_count = 0
        type_counts = Counter()
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty_count += 1
                    type_counts[cell.data_type] += 1
                    formula_count += int(cell.data_type == "f")
        workbook_meta.append(
            {
                "title": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": nonempty_count,
                "formula_cells": formula_count,
                "cell_data_types": dict(type_counts),
                "tables": list(sheet.tables.keys()),
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "auto_filter": sheet.auto_filter.ref,
                "hidden_rows": [idx for idx, dim in sheet.row_dimensions.items() if dim.hidden],
                "hidden_columns": [idx for idx, dim in sheet.column_dimensions.items() if dim.hidden],
                "comments": sum(
                    1 for row in sheet.iter_rows() for cell in row if cell.comment is not None
                ),
                "hyperlinks": sum(
                    1 for row in sheet.iter_rows() for cell in row if cell.hyperlink is not None
                ),
                "data_validations": len(sheet.data_validations.dataValidation),
            }
        )

    sheets = pd.read_excel(INPUT, sheet_name=None)
    summary = {"workbook": workbook_meta, "sheets": {}}
    for sheet_name, frame in sheets.items():
        columns = []
        for column in frame.columns:
            series = frame[column]
            python_types = Counter(type(value).__name__ for value in series.dropna())
            top_values = [
                {"value": normalise_scalar(value), "count": int(count)}
                for value, count in series.value_counts(dropna=False).head(20).items()
            ]
            columns.append(
                {
                    "name": str(column),
                    "pandas_dtype": str(series.dtype),
                    "non_null": int(series.notna().sum()),
                    "null": int(series.isna().sum()),
                    "unique_non_null": int(series.nunique(dropna=True)),
                    "python_types": dict(python_types),
                    "top_values": top_values,
                }
            )
        summary["sheets"][sheet_name] = {
            "rows": int(len(frame)),
            "columns": int(len(frame.columns)),
            "column_names": [str(column) for column in frame.columns],
            "column_profiles": columns,
        }

    main_frame = sheets["OxREP Mines 3 0 - 20250408"]
    main_frame.columns = main_frame.columns.astype(str)
    country_norm = main_frame["country"].astype("string").str.strip().str.casefold()
    spain = main_frame[country_norm.eq("spain")].copy()
    metal_columns = [column for column in main_frame if column.startswith("metalMined")]
    technique_columns = [column for column in main_frame if column.startswith("technique")]
    summary["spain_probe"] = {
        "rows_country_equals_spain": int(len(spain)),
        "mine_id_unique": int(spain["mineID"].nunique(dropna=True)),
        "duplicate_mine_id_rows": int(spain["mineID"].duplicated(keep=False).sum()),
        "exact_duplicate_rows": int(spain.duplicated(keep=False).sum()),
        "country_counts": {
            str(value): int(count)
            for value, count in main_frame["country"].value_counts(dropna=False).items()
        },
        "metal_columns": metal_columns,
        "technique_columns": technique_columns,
        "metal_raw_value_counts": {
            column: {
                str(value): int(count)
                for value, count in spain[column].value_counts(dropna=False).items()
            }
            for column in metal_columns
        },
        "technique_raw_value_counts": {
            column: {
                str(value): int(count)
                for value, count in spain[column].value_counts(dropna=False).items()
            }
            for column in technique_columns
        },
        "coordinate_ranges": {
            column: {
                "min": normalise_scalar(pd.to_numeric(spain[column], errors="coerce").min()),
                "max": normalise_scalar(pd.to_numeric(spain[column], errors="coerce").max()),
                "missing_or_non_numeric": int(
                    pd.to_numeric(spain[column], errors="coerce").isna().sum()
                ),
            }
            for column in ["latitude", "longitude", "coordinateAccuracy"]
        },
    }

    output = ROOT / "reports" / "audit" / "probe_excel.json"
    output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary["spain_probe"], ensure_ascii=False, indent=2))
    print("\nColumns:")
    for idx, column in enumerate(main_frame.columns, start=1):
        print(f"{idx:02d}. {column}")
    print(f"\nSaved: {output}")


if __name__ == "__main__":
    main()
