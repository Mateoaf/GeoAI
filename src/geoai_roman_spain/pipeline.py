from __future__ import annotations

import hashlib
import json
import math
import os
import platform
import re
import sqlite3
import sys
import unicodedata
import zipfile
from collections import Counter, OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


METALS = OrderedDict(
    [
        ("Au", ("Gold", "metalMinedGold")),
        ("Ag", ("Silver", "metalMinedSilver")),
        ("Pb", ("Lead", "metalMinedLead")),
        ("Cu", ("Copper", "metalMinedCopper")),
        ("Sn", ("Tin", "metalMinedTin")),
        ("Fe", ("Iron", "metalMinedIron")),
        ("Hg", ("Mercury/cinnabar", "metalMinedMercuryCinnabar")),
        ("Zn", ("Zinc", "metalMinedZinc")),
        ("Other", ("Other", "metalMinedOther")),
    ]
)

TECHNIQUES = OrderedDict(
    [
        ("opencast", ("Opencast / surface", "techniqueOpencast")),
        ("underground", ("Underground", "techniqueUnderground")),
        ("hydraulic", ("Hydraulic", "techniqueHydraulic")),
        ("hushing", ("Hushing", "techniqueHushing")),
        ("ground_sluicing", ("Ground sluicing", "techniqueGroundSluicing")),
        ("ruina_montium", ("Ruina montium", "techniqueRuinaMontium")),
        ("rake_comb", ("Rake / comb", "techniqueRakeComb")),
        ("gold_washing", ("Gold washing", "techniqueGoldWashing")),
        ("other", ("Other", "techniqueOther")),
    ]
)

CHRONOLOGY_COLUMNS = [
    "notBeforeOpeningDate",
    "notAfterOpeningDate",
    "notBeforeClosingDate",
    "notAfterClosingDate",
]

EXPECTED_DEPOSIT_TYPES = {
    "primary": "Primary",
    "secondary": "Secondary",
    "primary - secondary": "Primary + Secondary",
    "secondary - primary": "Primary + Secondary",
}

EXPECTED_EXPLOITATION_TYPES = {
    "selective": "Selective",
    "extensive": "Extensive",
    "selective - extensive": "Selective + Extensive",
    "extensive - selective": "Selective + Extensive",
}

TEXT_ARTIFACT_PATTERN = re.compile(
    r"(?:\ufffd|\u01fe|\u4000|_x[0-9A-Fa-f]{4}_)", re.IGNORECASE
)

QUALITY_FLAG_DEFINITIONS = OrderedDict(
    [
        (
            "flag_country_override",
            (
                "review",
                "Included by an explicit mineID override because the OxREP country field contains a subnational Spanish territory rather than 'Spain'.",
            ),
        ),
        (
            "flag_coordinate_missing",
            ("high", "Latitude or longitude is missing/non-numeric; no point geometry is exported."),
        ),
        (
            "flag_coordinate_outside_broad_spain_envelope",
            (
                "high",
                "Coordinate falls outside broad QA envelopes for Spain. These envelopes are not an administrative point-in-polygon test.",
            ),
        ),
        (
            "flag_coordinate_accuracy_missing",
            ("medium", "OxREP coordinateAccuracy is missing/non-numeric."),
        ),
        (
            "flag_coordinate_accuracy_zero",
            (
                "review",
                "OxREP reports coordinateAccuracy=0. Its semantics are not inferred as either perfect precision or missingness.",
            ),
        ),
        (
            "flag_coordinate_accuracy_gt_1000",
            (
                "medium",
                "Reported coordinateAccuracy is greater than 1000 (unit is not encoded in the worksheet schema).",
            ),
        ),
        (
            "flag_location_source_missing",
            ("medium", "locationDataSource is empty."),
        ),
        (
            "flag_metal_unknown_or_invalid",
            ("medium", "At least one commodity indicator is '?' or an unrecognised value."),
        ),
        (
            "flag_metal_missing",
            ("medium", "At least one commodity indicator is blank."),
        ),
        (
            "flag_no_confirmed_commodity",
            ("medium", "No commodity indicator is confirmed present."),
        ),
        (
            "flag_technique_unknown_or_invalid",
            ("medium", "At least one technique indicator is '?' or an unrecognised value."),
        ),
        (
            "flag_technique_incomplete",
            ("review", "At least one technique field is blank; blank is not treated as FALSE."),
        ),
        (
            "flag_no_confirmed_technique",
            ("review", "No technique indicator is confirmed present."),
        ),
        (
            "flag_chronology_missing",
            ("review", "All four chronology bounds and inUseDate are empty."),
        ),
        (
            "flag_chronology_invalid",
            ("high", "Available numeric chronology bounds are internally inconsistent."),
        ),
        (
            "flag_geology_missing",
            ("review", "geology is empty."),
        ),
        (
            "flag_deposit_type_missing",
            ("review", "depositType is empty."),
        ),
        (
            "flag_deposit_type_unexpected",
            ("medium", "depositType is populated but outside the observed Primary/Secondary vocabulary."),
        ),
        (
            "flag_exploitation_type_missing",
            ("review", "exploitationType is empty."),
        ),
        (
            "flag_exploitation_type_unexpected",
            ("medium", "exploitationType is populated but outside the observed Selective/Extensive vocabulary."),
        ),
        (
            "flag_possible_column_shift",
            (
                "high",
                "Values resemble a one-column shift between geology, depositType and exploitationType; requires source review.",
            ),
        ),
        (
            "flag_source_uncertainty_text",
            (
                "review",
                "Source text contains an explicit uncertainty/possible-duplicate/work-in-progress phrase.",
            ),
        ),
        (
            "flag_duplicate_mine_id",
            ("high", "mineID is duplicated within the selected dataset."),
        ),
        (
            "flag_duplicate_coordinates",
            ("review", "Another selected record has exactly the same numeric latitude and longitude."),
        ),
        (
            "flag_duplicate_site_name",
            ("review", "Another selected record has the same accent-insensitive normalised site name."),
        ),
        (
            "flag_nearby_record_100m",
            ("review", "Another selected record lies within the configured 100 m candidate-review threshold."),
        ),
        (
            "flag_text_encoding_artifact",
            ("medium", "At least one source text field contains a replacement/control artefact pattern."),
        ),
    ]
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def json_value(value: Any) -> Any:
    if value is pd.NA or value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def write_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(
        path,
        index=False,
        encoding="utf-8",
        lineterminator="\n",
        float_format="%.10g",
    )


def clean_text(value: Any) -> Any:
    if value is None or value is pd.NA:
        return pd.NA
    try:
        if pd.isna(value):
            return pd.NA
    except (TypeError, ValueError):
        pass
    text = unicodedata.normalize("NFC", str(value))
    text = re.sub(r"\s+", " ", text).strip()
    return text if text else pd.NA


def site_key(value: Any) -> str:
    cleaned = clean_text(value)
    if cleaned is pd.NA:
        return ""
    normalised = unicodedata.normalize("NFKD", str(cleaned))
    ascii_like = "".join(char for char in normalised if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", ascii_like.casefold())


def display_value(value: Any, limit: int = 240) -> str:
    converted = json_value(value)
    if converted is None:
        return "<NA>"
    text = str(converted)
    return text if len(text) <= limit else text[: limit - 1] + "…"


def semantic_role(column: str) -> str:
    if column == "mineID" or column.endswith("ID") or column == "spainMapSheet":
        return "identifier"
    if column in {"latitude", "longitude", "lambertX", "lambertY", "coordinateAccuracy"}:
        return "coordinate_or_precision"
    if column in CHRONOLOGY_COLUMNS or column == "inUseDate":
        return "chronology"
    if column.startswith("metalMined"):
        return "commodity_indicator"
    if column.startswith("technique"):
        return "technique_indicator"
    return "categorical_or_text"


def inspect_workbook(
    workbook_path: Path,
    report_path: Path,
    source_sheet: str,
    audit_dir: Path,
    methodological_references: list[dict[str, Any]],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    sheet_inventory: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        cell_types: Counter[str] = Counter()
        formula_count = 0
        nonempty_count = 0
        styled_nonempty_count = 0
        comments = 0
        hyperlinks = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty_count += 1
                    cell_types[cell.data_type] += 1
                    formula_count += int(cell.data_type == "f")
                    styled_nonempty_count += int(cell.has_style)
                comments += int(cell.comment is not None)
                hyperlinks += int(cell.hyperlink is not None)
        table_items = []
        for table in sheet.tables.values():
            table_items.append(
                {
                    "name": table.name,
                    "display_name": table.displayName,
                    "range": table.ref,
                    "style": table.tableStyleInfo.name if table.tableStyleInfo else None,
                }
            )
        sheet_inventory.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "dimension": sheet.calculate_dimension(),
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": nonempty_count,
                "styled_nonempty_cells": styled_nonempty_count,
                "formula_cells": formula_count,
                "cell_data_types": dict(sorted(cell_types.items())),
                "tables": table_items,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "auto_filter_range": sheet.auto_filter.ref,
                "hidden_rows": [idx for idx, dim in sheet.row_dimensions.items() if dim.hidden],
                "hidden_columns": [
                    idx for idx, dim in sheet.column_dimensions.items() if dim.hidden
                ],
                "comments": comments,
                "hyperlinks": hyperlinks,
                "data_validations": len(sheet.data_validations.dataValidation),
            }
        )

    defined_names = []
    for item in workbook.defined_names.values():
        defined_names.append(
            {
                "name": item.name,
                "formula": item.attr_text,
                "local_sheet_id": item.localSheetId,
            }
        )

    frames = pd.read_excel(workbook_path, sheet_name=None)
    if source_sheet not in frames:
        raise KeyError(f"Configured source sheet not found: {source_sheet}")
    source = frames[source_sheet].copy()
    source.columns = source.columns.astype(str)

    blank_cell_audit: dict[str, Any] = {}
    with zipfile.ZipFile(workbook_path) as archive:
        shared_string_count = 0
        empty_shared_strings = 0
        shared_strings_path = "xl/sharedStrings.xml"
        if shared_strings_path in archive.namelist():
            from xml.etree import ElementTree as ET

            shared_root = ET.fromstring(archive.read(shared_strings_path))
            namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            shared_items = shared_root.findall("x:si", namespace)
            shared_string_count = len(shared_items)
            empty_shared_strings = sum(
                1
                for item in shared_items
                if "".join(node.text or "" for node in item.findall(".//x:t", namespace)) == ""
            )
        blank_cell_audit = {
            "shared_string_count": shared_string_count,
            "empty_shared_string_entries": empty_shared_strings,
            "note": "An empty shared-string entry may be represented internally by its integer index; readers must resolve sharedStrings rather than treating the index as a cell value.",
        }

    reference_inventory = [
        {
            "file": item["path"].name,
            "relative_path": item["path"].relative_to(workbook_path.parents[2]).as_posix(),
            "role": item["role"],
            "sha256": sha256_file(item["path"]),
            "size_bytes": item["path"].stat().st_size,
        }
        for item in methodological_references
    ]
    workbook_inventory = {
        "workbook_file": workbook_path.name,
        "workbook_sha256": sha256_file(workbook_path),
        "workbook_size_bytes": workbook_path.stat().st_size,
        "report_file": report_path.name,
        "report_sha256": sha256_file(report_path),
        "report_size_bytes": report_path.stat().st_size,
        "methodological_references": reference_inventory,
        "sheet_count": len(workbook.sheetnames),
        "sheet_order": workbook.sheetnames,
        "active_sheet": workbook.active.title,
        "defined_names": defined_names,
        "sheets": sheet_inventory,
        "pandas_shapes": {
            name: {"rows": int(frame.shape[0]), "columns": int(frame.shape[1])}
            for name, frame in frames.items()
        },
        "blank_cell_audit": blank_cell_audit,
    }
    write_json(audit_dir / "workbook_inventory.json", workbook_inventory)

    column_rows: list[dict[str, Any]] = []
    value_count_rows: list[dict[str, Any]] = []
    for index, column in enumerate(source.columns, start=1):
        series = source[column]
        non_null = series.dropna()
        numeric = pd.to_numeric(series, errors="coerce")
        type_counts = Counter(type(value).__name__ for value in non_null)
        unique_examples = [display_value(value) for value in non_null.unique()[:8]]
        text_lengths = non_null.astype(str).str.len() if len(non_null) else pd.Series(dtype=int)
        column_rows.append(
            {
                "column_index": index,
                "excel_column": get_column_letter(index),
                "column": column,
                "semantic_role": semantic_role(column),
                "pandas_dtype": str(series.dtype),
                "observed_python_types": json.dumps(dict(type_counts), sort_keys=True),
                "rows": int(len(series)),
                "non_null": int(series.notna().sum()),
                "null": int(series.isna().sum()),
                "null_pct": float(series.isna().mean()),
                "unique_non_null": int(series.nunique(dropna=True)),
                "numeric_parse_success_non_null": int(numeric.notna().sum()),
                "numeric_parse_ratio_non_null": (
                    float(numeric.notna().sum() / len(non_null)) if len(non_null) else None
                ),
                "numeric_min": json_value(numeric.min()),
                "numeric_max": json_value(numeric.max()),
                "max_text_length": int(text_lengths.max()) if len(text_lengths) else 0,
                "examples": json.dumps(unique_examples, ensure_ascii=False),
            }
        )
        counts = series.value_counts(dropna=False, sort=True)
        for rank, (value, count) in enumerate(counts.items(), start=1):
            value_count_rows.append(
                {
                    "column": column,
                    "rank_within_column": rank,
                    "value": display_value(value, limit=2000),
                    "value_type": "missing" if json_value(value) is None else type(value).__name__,
                    "count": int(count),
                    "pct_rows": float(count / len(source)) if len(source) else None,
                }
            )

    write_csv(pd.DataFrame(column_rows), audit_dir / "column_audit.csv")
    write_csv(pd.DataFrame(value_count_rows), audit_dir / "value_counts_all_columns.csv")
    write_json(
        audit_dir / "input_manifest.json",
        {
            "files": [
                {
                    "role": "OxREP source workbook (immutable project copy)",
                    "relative_path": str(workbook_path.relative_to(workbook_path.parents[2])).replace("\\", "/"),
                    "size_bytes": workbook_path.stat().st_size,
                    "sha256": sha256_file(workbook_path),
                },
                {
                    "role": "Project concept report (immutable project copy)",
                    "relative_path": str(report_path.relative_to(report_path.parents[2])).replace("\\", "/"),
                    "size_bytes": report_path.stat().st_size,
                    "sha256": sha256_file(report_path),
                },
            ]
            + [
                {
                    "role": item["role"],
                    "relative_path": item["path"].relative_to(workbook_path.parents[2]).as_posix(),
                    "size_bytes": item["path"].stat().st_size,
                    "sha256": sha256_file(item["path"]),
                }
                for item in methodological_references
            ]
        },
    )
    return source, workbook_inventory


def parse_indicator(value: Any) -> tuple[Any, str]:
    """Map heterogeneous OxREP indicator values without conflating missing/unknown/false."""
    if value is None or value is pd.NA:
        return pd.NA, "missing"
    try:
        if pd.isna(value):
            return pd.NA, "missing"
    except (TypeError, ValueError):
        pass
    if isinstance(value, (bool, np.bool_)):
        return bool(value), "present" if bool(value) else "absent"
    if isinstance(value, (int, float, np.integer, np.floating)):
        if float(value) == 1.0:
            return True, "present"
        if float(value) == 0.0:
            return False, "absent"
        return pd.NA, "invalid"
    text = str(value).strip().casefold()
    if text in {"true", "1", "yes", "y"}:
        return True, "present"
    if text in {"false", "0", "no", "n"}:
        return False, "absent"
    if text in {"?", "unknown", "uncertain"}:
        return pd.NA, "unknown"
    if not text:
        return pd.NA, "missing"
    return pd.NA, "invalid"


def in_broad_spain_envelope(
    latitude: Any, longitude: Any, envelopes: Iterable[dict[str, float]]
) -> bool:
    try:
        lat = float(latitude)
        lon = float(longitude)
    except (TypeError, ValueError):
        return False
    if not math.isfinite(lat) or not math.isfinite(lon):
        return False
    return any(
        envelope["min_lat"] <= lat <= envelope["max_lat"]
        and envelope["min_lon"] <= lon <= envelope["max_lon"]
        for envelope in envelopes
    )


def row_fingerprint(row: pd.Series, columns: list[str]) -> str:
    values = [json_value(row[column]) for column in columns]
    encoded = json.dumps(values, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def make_selection_decisions(
    source: pd.DataFrame, config: dict[str, Any]
) -> pd.DataFrame:
    selection = config["selection"]
    primary = selection["primary_country"].strip().casefold()
    override_ids = {int(item) for item in selection["manual_include_mine_ids"]}
    expected_override_country = selection["manual_include_expected_country"].strip().casefold()
    envelopes = config["coordinate_plausibility_envelopes"]
    rows: list[dict[str, Any]] = []
    seen_override_ids: set[int] = set()
    for source_index, row in source.iterrows():
        mine_id_numeric = pd.to_numeric(pd.Series([row.get("mineID")]), errors="coerce").iloc[0]
        mine_id = int(mine_id_numeric) if pd.notna(mine_id_numeric) else None
        country_value = clean_text(row.get("country"))
        country_key = "" if country_value is pd.NA else str(country_value).casefold()
        coordinate_plausible = in_broad_spain_envelope(
            row.get("latitude"), row.get("longitude"), envelopes
        )
        include = False
        rule = "excluded_country_not_spain"
        reason = "country is not the configured modern-country value Spain."
        if country_key == primary:
            include = True
            rule = "included_country_exact"
            reason = "Trimmed, case-insensitive country value equals Spain."
        elif mine_id in override_ids:
            if country_key != expected_override_country:
                raise ValueError(
                    f"Configured override mineID {mine_id} no longer has expected country "
                    f"{selection['manual_include_expected_country']!r}."
                )
            if not coordinate_plausible:
                raise ValueError(
                    f"Configured override mineID {mine_id} falls outside broad Spain QA envelopes."
                )
            include = True
            seen_override_ids.add(mine_id)
            rule = "included_reviewed_override"
            reason = selection["manual_include_reason"]
        rows.append(
            {
                "source_excel_row": int(source_index) + 2,
                "mineID": mine_id,
                "site": json_value(row.get("site")),
                "country_raw": json_value(row.get("country")),
                "roman_province_raw": json_value(row.get("province")),
                "region_raw": json_value(row.get("region")),
                "latitude_raw": json_value(row.get("latitude")),
                "longitude_raw": json_value(row.get("longitude")),
                "coordinate_in_broad_spain_envelope": coordinate_plausible,
                "include_spain_subset": include,
                "selection_rule": rule,
                "selection_reason": reason,
            }
        )
    if seen_override_ids != override_ids:
        missing = sorted(override_ids - seen_override_ids)
        raise ValueError(f"Configured Spain overrides not found/reviewed: {missing}")
    return pd.DataFrame(rows)


def haversine_metres(lat1: float, lon1: float, lat2: np.ndarray, lon2: np.ndarray) -> np.ndarray:
    radius = 6_371_008.8
    phi1 = math.radians(lat1)
    phi2 = np.radians(lat2)
    dphi = phi2 - phi1
    dlambda = np.radians(lon2) - math.radians(lon1)
    a = np.sin(dphi / 2.0) ** 2 + math.cos(phi1) * np.cos(phi2) * np.sin(dlambda / 2.0) ** 2
    return radius * 2.0 * np.arctan2(np.sqrt(a), np.sqrt(1.0 - a))


def duplicate_candidates(
    frame: pd.DataFrame, threshold_m: float
) -> tuple[pd.DataFrame, set[int], set[int], set[int]]:
    pair_reasons: dict[tuple[int, int], set[str]] = {}
    pair_distances: dict[tuple[int, int], float | None] = {}

    site_groups = frame.groupby("site_match_key", dropna=False).groups
    duplicate_site_rows: set[int] = set()
    for key, indexes in site_groups.items():
        index_list = list(indexes)
        if not key or len(index_list) < 2:
            continue
        duplicate_site_rows.update(index_list)
        for pos, left in enumerate(index_list[:-1]):
            for right in index_list[pos + 1 :]:
                pair = (min(left, right), max(left, right))
                pair_reasons.setdefault(pair, set()).add("same_normalised_site")
                pair_distances.setdefault(pair, None)

    coords = frame[
        frame["latitude_decimal"].notna() & frame["longitude_decimal"].notna()
    ][["latitude_decimal", "longitude_decimal"]].copy()
    coord_groups = coords.groupby(["latitude_decimal", "longitude_decimal"]).groups
    duplicate_coordinate_rows: set[int] = set()
    for indexes in coord_groups.values():
        index_list = list(indexes)
        if len(index_list) < 2:
            continue
        duplicate_coordinate_rows.update(index_list)
        for pos, left in enumerate(index_list[:-1]):
            for right in index_list[pos + 1 :]:
                pair = (min(left, right), max(left, right))
                pair_reasons.setdefault(pair, set()).add("exact_same_coordinates")
                pair_distances[pair] = 0.0

    valid_indexes = coords.index.to_numpy()
    latitudes = coords["latitude_decimal"].to_numpy(dtype=float)
    longitudes = coords["longitude_decimal"].to_numpy(dtype=float)
    nearby_rows: set[int] = set()
    for pos in range(len(coords) - 1):
        distances = haversine_metres(
            latitudes[pos],
            longitudes[pos],
            latitudes[pos + 1 :],
            longitudes[pos + 1 :],
        )
        matches = np.flatnonzero(distances <= threshold_m)
        for match in matches:
            left = int(valid_indexes[pos])
            right = int(valid_indexes[pos + 1 + match])
            pair = (min(left, right), max(left, right))
            pair_reasons.setdefault(pair, set()).add(f"within_{threshold_m:g}_m")
            pair_distances[pair] = float(distances[match])
            nearby_rows.update(pair)

    candidate_rows: list[dict[str, Any]] = []
    for pair in sorted(pair_reasons):
        left_index, right_index = pair
        left = frame.loc[left_index]
        right = frame.loc[right_index]
        distance = pair_distances.get(pair)
        if distance is None and all(
            pd.notna(value)
            for value in [
                left["latitude_decimal"],
                left["longitude_decimal"],
                right["latitude_decimal"],
                right["longitude_decimal"],
            ]
        ):
            distance = float(
                haversine_metres(
                    float(left["latitude_decimal"]),
                    float(left["longitude_decimal"]),
                    np.array([float(right["latitude_decimal"])]),
                    np.array([float(right["longitude_decimal"])]),
                )[0]
            )
        candidate_rows.append(
            {
                "mineID_left": int(left["mineID"]),
                "mineID_right": int(right["mineID"]),
                "site_left": left["site"],
                "site_right": right["site"],
                "latitude_left": json_value(left["latitude_decimal"]),
                "longitude_left": json_value(left["longitude_decimal"]),
                "latitude_right": json_value(right["latitude_decimal"]),
                "longitude_right": json_value(right["longitude_decimal"]),
                "distance_m": distance,
                "candidate_reasons": ";".join(sorted(pair_reasons[pair])),
                "decision": "retain_both_pending_domain_review",
            }
        )
    return (
        pd.DataFrame(candidate_rows),
        duplicate_site_rows,
        duplicate_coordinate_rows,
        nearby_rows,
    )


def chronology_invalid(row: pd.Series) -> bool:
    opening_low = row.get("notBeforeOpeningDate_numeric")
    opening_high = row.get("notAfterOpeningDate_numeric")
    closing_low = row.get("notBeforeClosingDate_numeric")
    closing_high = row.get("notAfterClosingDate_numeric")
    if pd.notna(opening_low) and pd.notna(opening_high) and opening_low > opening_high:
        return True
    if pd.notna(closing_low) and pd.notna(closing_high) and closing_low > closing_high:
        return True
    if pd.notna(opening_low) and pd.notna(closing_high) and opening_low > closing_high:
        return True
    return False


def prepare_spain_dataset(
    source: pd.DataFrame,
    config: dict[str, Any],
    workbook_sha256: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    original_columns = list(source.columns)
    decisions = make_selection_decisions(source, config)
    selected_rows = decisions.loc[decisions["include_spain_subset"], "source_excel_row"] - 2
    clean = source.loc[selected_rows.tolist()].copy()
    clean.insert(0, "source_excel_row", clean.index.astype(int) + 2)
    clean.insert(1, "source_sheet", config["source_sheet"])
    clean.insert(2, "source_file", Path(config["input_workbook"]).name)
    clean.insert(3, "source_file_sha256", workbook_sha256)
    clean.insert(
        4,
        "source_row_sha256",
        clean.apply(lambda row: row_fingerprint(row, original_columns), axis=1),
    )
    clean.reset_index(drop=True, inplace=True)
    clean["mineID"] = pd.to_numeric(clean["mineID"], errors="coerce").astype("Int64")

    text_derivatives = {
        "site_clean": "site",
        "ancient_name_clean": "ancientName",
        "roman_province_clean": "province",
        "country_clean": "country",
        "region_clean": "region",
        "mining_district_clean": "miningDistrict",
        "in_use_date_clean": "inUseDate",
        "geology_clean": "geology",
        "deposit_type_clean": "depositType",
        "exploitation_type_clean": "exploitationType",
    }
    for target, source_column in text_derivatives.items():
        clean[target] = clean[source_column].map(clean_text).astype("string")
    clean["site_match_key"] = clean["site"].map(site_key).astype("string")

    clean["latitude_decimal"] = pd.to_numeric(clean["latitude"], errors="coerce")
    clean["longitude_decimal"] = pd.to_numeric(clean["longitude"], errors="coerce")
    clean["coordinate_accuracy_reported"] = pd.to_numeric(
        clean["coordinateAccuracy"], errors="coerce"
    )
    clean["coordinate_in_broad_spain_envelope"] = clean.apply(
        lambda row: in_broad_spain_envelope(
            row["latitude_decimal"],
            row["longitude_decimal"],
            config["coordinate_plausibility_envelopes"],
        ),
        axis=1,
    )

    commodity_rows: list[dict[str, Any]] = []
    for code, (name, source_column) in METALS.items():
        parsed = clean[source_column].map(parse_indicator)
        clean[f"Y_{code}"] = pd.array([item[0] for item in parsed], dtype="boolean")
        clean[f"commodity_{code}_status"] = pd.array(
            [item[1] for item in parsed], dtype="string"
        )
        for row_index, row in clean.iterrows():
            commodity_rows.append(
                {
                    "mineID": int(row["mineID"]),
                    "commodity_code": code,
                    "commodity_name": name,
                    "source_column": source_column,
                    "source_raw_value": json_value(row[source_column]),
                    "status": row[f"commodity_{code}_status"],
                    "confirmed_present": json_value(row[f"Y_{code}"]),
                }
            )
    clean["Y_roman"] = 1
    y_columns = [f"Y_{code}" for code in METALS]
    clean["confirmed_commodity_count"] = clean[y_columns].fillna(False).sum(axis=1).astype(int)
    clean["commodity_combination"] = clean.apply(
        lambda row: "+".join(
            code
            for code in METALS
            if pd.notna(row[f"Y_{code}"]) and bool(row[f"Y_{code}"])
        )
        or "None confirmed",
        axis=1,
    )
    clean["commodity_information_complete"] = clean[
        [f"commodity_{code}_status" for code in METALS]
    ].apply(lambda row: all(value in {"present", "absent"} for value in row), axis=1)

    technique_rows: list[dict[str, Any]] = []
    for key, (name, source_column) in TECHNIQUES.items():
        parsed = clean[source_column].map(parse_indicator)
        clean[f"has_technique_{key}"] = pd.array(
            [item[0] for item in parsed], dtype="boolean"
        )
        clean[f"technique_{key}_status"] = pd.array(
            [item[1] for item in parsed], dtype="string"
        )
        for row_index, row in clean.iterrows():
            technique_rows.append(
                {
                    "mineID": int(row["mineID"]),
                    "technique_code": key,
                    "technique_name": name,
                    "source_column": source_column,
                    "source_raw_value": json_value(row[source_column]),
                    "status": row[f"technique_{key}_status"],
                    "confirmed_present": json_value(row[f"has_technique_{key}"]),
                }
            )
    technique_boolean_columns = [f"has_technique_{key}" for key in TECHNIQUES]
    clean["confirmed_technique_count"] = (
        clean[technique_boolean_columns].fillna(False).sum(axis=1).astype(int)
    )
    clean["headline_technique_combination"] = clean.apply(
        lambda row: "+".join(
            key
            for key in ["opencast", "underground", "hydraulic"]
            if pd.notna(row[f"has_technique_{key}"])
            and bool(row[f"has_technique_{key}"])
        )
        or "None confirmed",
        axis=1,
    )

    for column in CHRONOLOGY_COLUMNS:
        clean[f"{column}_numeric"] = pd.to_numeric(clean[column], errors="coerce")
    clean["chronology_numeric_bounds_count"] = clean[
        [f"{column}_numeric" for column in CHRONOLOGY_COLUMNS]
    ].notna().sum(axis=1)
    clean["chronology_any_reported"] = (
        clean["chronology_numeric_bounds_count"].gt(0) | clean["inUseDate"].notna()
    )

    deposit_key = clean["deposit_type_clean"].str.casefold()
    exploitation_key = clean["exploitation_type_clean"].str.casefold()
    clean["deposit_type_normalized"] = deposit_key.map(EXPECTED_DEPOSIT_TYPES).astype("string")
    clean["exploitation_type_normalized"] = exploitation_key.map(
        EXPECTED_EXPLOITATION_TYPES
    ).astype("string")

    candidates, duplicate_sites, duplicate_coordinates, nearby_rows = duplicate_candidates(
        clean, float(config["duplicate_distance_threshold_m"])
    )

    clean["flag_country_override"] = clean["mineID"].isin(
        config["selection"]["manual_include_mine_ids"]
    )
    clean["flag_coordinate_missing"] = (
        clean["latitude_decimal"].isna() | clean["longitude_decimal"].isna()
    )
    clean["flag_coordinate_outside_broad_spain_envelope"] = (
        ~clean["flag_coordinate_missing"]
        & ~clean["coordinate_in_broad_spain_envelope"]
    )
    clean["flag_coordinate_accuracy_missing"] = clean[
        "coordinate_accuracy_reported"
    ].isna()
    clean["flag_coordinate_accuracy_zero"] = clean[
        "coordinate_accuracy_reported"
    ].eq(0)
    clean["flag_coordinate_accuracy_gt_1000"] = clean[
        "coordinate_accuracy_reported"
    ].gt(1000)
    clean["flag_location_source_missing"] = clean["locationDataSource"].isna()
    commodity_status_columns = [f"commodity_{code}_status" for code in METALS]
    clean["flag_metal_unknown_or_invalid"] = clean[commodity_status_columns].apply(
        lambda row: any(value in {"unknown", "invalid"} for value in row), axis=1
    )
    clean["flag_metal_missing"] = clean[commodity_status_columns].apply(
        lambda row: any(value == "missing" for value in row), axis=1
    )
    clean["flag_no_confirmed_commodity"] = clean["confirmed_commodity_count"].eq(0)
    technique_status_columns = [f"technique_{key}_status" for key in TECHNIQUES]
    clean["flag_technique_unknown_or_invalid"] = clean[technique_status_columns].apply(
        lambda row: any(value in {"unknown", "invalid"} for value in row), axis=1
    )
    clean["flag_technique_incomplete"] = clean[technique_status_columns].apply(
        lambda row: any(value == "missing" for value in row), axis=1
    )
    clean["flag_no_confirmed_technique"] = clean["confirmed_technique_count"].eq(0)
    clean["flag_chronology_missing"] = ~clean["chronology_any_reported"]
    clean["flag_chronology_invalid"] = clean.apply(chronology_invalid, axis=1)
    clean["flag_geology_missing"] = clean["geology"].isna()
    clean["flag_deposit_type_missing"] = clean["depositType"].isna()
    clean["flag_deposit_type_unexpected"] = (
        clean["depositType"].notna() & clean["deposit_type_normalized"].isna()
    )
    clean["flag_exploitation_type_missing"] = clean["exploitationType"].isna()
    clean["flag_exploitation_type_unexpected"] = (
        clean["exploitationType"].notna()
        & clean["exploitation_type_normalized"].isna()
    )
    clean["flag_possible_column_shift"] = (
        clean["deposit_type_clean"].str.casefold().isin({"hard rock"})
        | clean["exploitation_type_clean"].str.casefold().isin(
            {"primary", "secondary", "hard rock", "pyrite"}
        )
    ).fillna(False)
    uncertainty_columns = [
        "locationDataSource",
        "locationNotes",
        "description",
        "references",
        "notes",
    ]
    uncertainty_pattern = re.compile(
        r"probably same|possibly duplicate|possible duplicate|guessed|"
        r"need to add details|does not give coordinates|uncertain|perhaps|\?",
        re.IGNORECASE,
    )
    clean["flag_source_uncertainty_text"] = clean.apply(
        lambda row: any(
            isinstance(row.get(column), str)
            and uncertainty_pattern.search(row[column]) is not None
            for column in uncertainty_columns
        ),
        axis=1,
    )
    clean["flag_duplicate_mine_id"] = clean["mineID"].duplicated(keep=False)
    clean["flag_duplicate_coordinates"] = clean.index.isin(duplicate_coordinates)
    clean["flag_duplicate_site_name"] = clean.index.isin(duplicate_sites)
    clean["flag_nearby_record_100m"] = clean.index.isin(nearby_rows)
    source_text_columns = [
        column
        for column in original_columns
        if source[column].dtype == object or pd.api.types.is_string_dtype(source[column])
    ]
    clean["text_artifact_columns"] = clean.apply(
        lambda row: ";".join(
            column
            for column in source_text_columns
            if isinstance(row.get(column), str) and TEXT_ARTIFACT_PATTERN.search(row[column])
        ),
        axis=1,
    )
    clean["flag_text_encoding_artifact"] = clean["text_artifact_columns"].ne("")

    flag_columns = list(QUALITY_FLAG_DEFINITIONS)
    for column in flag_columns:
        clean[column] = clean[column].astype(bool)
    clean["quality_flag_count"] = clean[flag_columns].sum(axis=1).astype(int)
    high_flags = [
        column
        for column, (severity, _definition) in QUALITY_FLAG_DEFINITIONS.items()
        if severity == "high"
    ]
    clean["quality_high_severity"] = clean[high_flags].any(axis=1)
    clean["quality_review_required"] = clean[flag_columns].any(axis=1)

    quality_definitions = pd.DataFrame(
        [
            {"flag": flag, "severity": severity, "definition": definition}
            for flag, (severity, definition) in QUALITY_FLAG_DEFINITIONS.items()
        ]
    )
    return (
        clean,
        decisions,
        pd.DataFrame(commodity_rows),
        pd.DataFrame(technique_rows),
        candidates,
        quality_definitions,
    )


def count_table(
    series: pd.Series,
    category_name: str,
    total: int,
    include_missing: bool = True,
) -> pd.DataFrame:
    values = series.astype("object")
    if include_missing:
        values = values.where(pd.notna(values), "<Missing>")
    else:
        values = values[pd.notna(values)]
    counts = values.value_counts(dropna=False)
    return pd.DataFrame(
        {
            category_name: [json_value(value) for value in counts.index],
            "count": counts.values.astype(int),
            "pct_records": counts.values / total if total else np.nan,
        }
    )


def build_summary_tables(
    clean: pd.DataFrame,
    commodity_long: pd.DataFrame,
    technique_long: pd.DataFrame,
    reports_tables: Path,
) -> dict[str, pd.DataFrame]:
    reports_tables.mkdir(parents=True, exist_ok=True)
    total = len(clean)
    tables: dict[str, pd.DataFrame] = {}

    commodity_rows = []
    for code, (name, source_column) in METALS.items():
        status = clean[f"commodity_{code}_status"]
        commodity_rows.append(
            {
                "commodity_code": code,
                "commodity_name": name,
                "source_column": source_column,
                "confirmed_present": int(status.eq("present").sum()),
                "confirmed_absent": int(status.eq("absent").sum()),
                "unknown": int(status.eq("unknown").sum()),
                "missing": int(status.eq("missing").sum()),
                "invalid": int(status.eq("invalid").sum()),
                "prevalence_among_selected": float(status.eq("present").mean()),
            }
        )
    tables["commodity_summary"] = pd.DataFrame(commodity_rows)
    tables["commodity_combinations"] = count_table(
        clean["commodity_combination"], "commodity_combination", total, False
    )
    tables["commodity_count_distribution"] = count_table(
        clean["confirmed_commodity_count"], "confirmed_commodity_count", total, False
    )

    technique_rows = []
    for key, (name, source_column) in TECHNIQUES.items():
        status = clean[f"technique_{key}_status"]
        technique_rows.append(
            {
                "technique_code": key,
                "technique_name": name,
                "source_column": source_column,
                "confirmed_present": int(status.eq("present").sum()),
                "confirmed_absent": int(status.eq("absent").sum()),
                "unknown": int(status.eq("unknown").sum()),
                "missing": int(status.eq("missing").sum()),
                "invalid": int(status.eq("invalid").sum()),
                "prevalence_among_selected": float(status.eq("present").mean()),
            }
        )
    tables["technique_summary"] = pd.DataFrame(technique_rows)
    tables["headline_technique_combinations"] = count_table(
        clean["headline_technique_combination"],
        "headline_technique_combination",
        total,
        False,
    )

    tables["coordinate_accuracy"] = count_table(
        clean["coordinate_accuracy_reported"], "coordinate_accuracy_reported", total, True
    )
    tables["roman_province"] = count_table(
        clean["roman_province_clean"], "roman_province", total, True
    )
    tables["region"] = count_table(clean["region_clean"], "region", total, True)
    tables["mining_district"] = count_table(
        clean["mining_district_clean"], "mining_district", total, True
    )
    tables["geology"] = count_table(clean["geology_clean"], "geology", total, True)
    tables["deposit_type_raw"] = count_table(
        clean["deposit_type_clean"], "deposit_type_raw", total, True
    )
    tables["deposit_type_normalized"] = count_table(
        clean["deposit_type_normalized"], "deposit_type_normalized", total, True
    )
    tables["exploitation_type_raw"] = count_table(
        clean["exploitation_type_clean"], "exploitation_type_raw", total, True
    )
    tables["exploitation_type_normalized"] = count_table(
        clean["exploitation_type_normalized"],
        "exploitation_type_normalized",
        total,
        True,
    )
    tables["in_use_date"] = count_table(
        clean["in_use_date_clean"], "in_use_date", total, True
    )

    quality_rows = []
    for flag, (severity, definition) in QUALITY_FLAG_DEFINITIONS.items():
        count = int(clean[flag].sum())
        quality_rows.append(
            {
                "flag": flag,
                "severity": severity,
                "count": count,
                "pct_records": count / total if total else None,
                "definition": definition,
            }
        )
    tables["quality_flag_summary"] = pd.DataFrame(quality_rows)

    numeric_chronology_rows = []
    for column in CHRONOLOGY_COLUMNS:
        numeric = clean[f"{column}_numeric"]
        numeric_chronology_rows.append(
            {
                "source_column": column,
                "non_null_numeric": int(numeric.notna().sum()),
                "missing_or_non_numeric": int(numeric.isna().sum()),
                "min": json_value(numeric.min()),
                "median": json_value(numeric.median()),
                "max": json_value(numeric.max()),
            }
        )
    tables["chronology_numeric_summary"] = pd.DataFrame(numeric_chronology_rows)

    commodity_region_rows = []
    for region, subset in clean.groupby("region_clean", dropna=False):
        for code in METALS:
            commodity_region_rows.append(
                {
                    "region": json_value(region) or "<Missing>",
                    "commodity_code": code,
                    "records": int(len(subset)),
                    "confirmed_present": int(subset[f"Y_{code}"].fillna(False).sum()),
                }
            )
    tables["commodity_by_region"] = pd.DataFrame(commodity_region_rows)

    for name, frame in tables.items():
        write_csv(frame, reports_tables / f"{name}.csv")
    return tables


def dataframe_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return [
        {column: json_value(value) for column, value in row.items()}
        for row in frame.to_dict(orient="records")
    ]


def export_geojson(clean: pd.DataFrame, path: Path) -> None:
    features = []
    for _, row in clean.iterrows():
        if pd.isna(row["latitude_decimal"]) or pd.isna(row["longitude_decimal"]):
            continue
        properties = {
            "mineID": int(row["mineID"]),
            "site": json_value(row["site_clean"]),
            "country_raw": json_value(row["country"]),
            "roman_province": json_value(row["roman_province_clean"]),
            "region": json_value(row["region_clean"]),
            "mining_district": json_value(row["mining_district_clean"]),
            "coordinate_accuracy_reported": json_value(
                row["coordinate_accuracy_reported"]
            ),
            "commodity_combination": row["commodity_combination"],
            "quality_flag_count": int(row["quality_flag_count"]),
            "source_excel_row": int(row["source_excel_row"]),
        }
        properties.update(
            {f"Y_{code}": json_value(row[f"Y_{code}"]) for code in METALS}
        )
        properties.update(
            {
                f"technique_{key}": json_value(row[f"has_technique_{key}"])
                for key in TECHNIQUES
            }
        )
        features.append(
            {
                "type": "Feature",
                "id": int(row["mineID"]),
                "properties": properties,
                "geometry": {
                    "type": "Point",
                    "coordinates": [
                        float(row["longitude_decimal"]),
                        float(row["latitude_decimal"]),
                    ],
                },
            }
        )
    write_json(
        path,
        {
            "type": "FeatureCollection",
            "name": "oxrep_roman_mines_spain_clean",
            "crs": {
                "type": "name",
                "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"},
            },
            "features": features,
        },
    )


def export_sqlite(
    clean: pd.DataFrame,
    commodity_long: pd.DataFrame,
    technique_long: pd.DataFrame,
    quality_definitions: pd.DataFrame,
    duplicate_candidates_frame: pd.DataFrame,
    path: Path,
) -> None:
    if path.exists():
        path.unlink()
    mine_columns = [
        "mineID",
        "source_excel_row",
        "source_sheet",
        "source_file",
        "source_file_sha256",
        "source_row_sha256",
        "site",
        "ancientName",
        "province",
        "country",
        "region",
        "miningDistrict",
        "latitude_decimal",
        "longitude_decimal",
        "coordinate_accuracy_reported",
        "locationDataSource",
        "locationNotes",
        "description",
        "references",
        "notes",
        "geology",
        "depositType",
        "exploitationType",
        "commodity_combination",
        "confirmed_commodity_count",
        "confirmed_technique_count",
        "quality_flag_count",
    ] + list(QUALITY_FLAG_DEFINITIONS)
    mines = clean[mine_columns].copy()
    for column in mines.select_dtypes(include=["boolean", "bool"]).columns:
        mines[column] = mines[column].astype("Int64")
    flags_long_rows = []
    for _, row in clean.iterrows():
        for flag, (severity, definition) in QUALITY_FLAG_DEFINITIONS.items():
            if bool(row[flag]):
                flags_long_rows.append(
                    {
                        "mineID": int(row["mineID"]),
                        "flag": flag,
                        "severity": severity,
                        "definition": definition,
                    }
                )
    with sqlite3.connect(path) as connection:
        mines.to_sql("roman_mine", connection, index=False, if_exists="replace")
        commodity_long.to_sql("mine_commodity", connection, index=False, if_exists="replace")
        technique_long.to_sql("mine_technique", connection, index=False, if_exists="replace")
        pd.DataFrame(flags_long_rows).to_sql(
            "mine_quality_flag", connection, index=False, if_exists="replace"
        )
        quality_definitions.to_sql(
            "quality_flag_definition", connection, index=False, if_exists="replace"
        )
        duplicate_candidates_frame.to_sql(
            "duplicate_candidate", connection, index=False, if_exists="replace"
        )
        connection.executescript(
            """
            CREATE UNIQUE INDEX idx_roman_mine_mineID ON roman_mine(mineID);
            CREATE INDEX idx_mine_commodity_mineID ON mine_commodity(mineID);
            CREATE INDEX idx_mine_technique_mineID ON mine_technique(mineID);
            CREATE INDEX idx_quality_flag_mineID ON mine_quality_flag(mineID);
            """
        )


def draw_title(draw: ImageDraw.ImageDraw, title: str, subtitle: str, width: int) -> int:
    title_font = ImageFont.truetype("arialbd.ttf", 42)
    subtitle_font = ImageFont.truetype("arial.ttf", 22)
    draw.text((70, 45), title, font=title_font, fill="#133C55")
    draw.text((70, 105), subtitle, font=subtitle_font, fill="#52606D")
    draw.line((70, 145, width - 70, 145), fill="#CBD5E1", width=2)
    return 175


def draw_bar_chart(
    labels: list[str],
    values: list[int],
    title: str,
    subtitle: str,
    path: Path,
    color: str = "#1B7F79",
    footnote: str | None = None,
) -> None:
    width = 1500
    row_height = 48
    footer_space = 120 if footnote else 35
    height = 220 + row_height * max(len(labels), 1) + footer_space
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    y = draw_title(draw, title, subtitle, width)
    label_font = ImageFont.truetype("arial.ttf", 20)
    value_font = ImageFont.truetype("arialbd.ttf", 20)
    note_font = ImageFont.truetype("arial.ttf", 18)
    label_width = 390
    chart_left = 480
    chart_right = width - 120
    max_value = max(values) if values else 1
    for label, value in zip(labels, values):
        y_mid = y + row_height // 2
        label_text = label if len(label) <= 34 else label[:33] + "…"
        draw.text((70, y_mid - 12), label_text, font=label_font, fill="#1F2933")
        bar_width = int((chart_right - chart_left) * value / max_value) if max_value else 0
        draw.rounded_rectangle(
            (chart_left, y_mid - 14, chart_left + max(bar_width, 2), y_mid + 14),
            radius=7,
            fill=color,
        )
        draw.text(
            (chart_left + max(bar_width, 2) + 12, y_mid - 12),
            f"{value:,}",
            font=value_font,
            fill="#102A43",
        )
        y += row_height
    if footnote:
        draw.text((70, height - 60), footnote, font=note_font, fill="#64748B")
    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path, optimize=True)


def draw_map(clean: pd.DataFrame, path: Path) -> None:
    valid = clean[
        clean["latitude_decimal"].notna() & clean["longitude_decimal"].notna()
    ].copy()
    width, height = 1500, 1050
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    top = draw_title(
        draw,
        "Distribución espacial de minas romanas seleccionadas",
        "Puntos OxREP; visualización exploratoria sin cartografía base ni inferencia geológica",
        width,
    )
    left, right = 100, width - 100
    bottom = height - 100
    min_lon, max_lon = -9.6, 4.6
    min_lat, max_lat = 35.4, 44.6

    def xy(lon: float, lat: float) -> tuple[float, float]:
        x = left + (lon - min_lon) / (max_lon - min_lon) * (right - left)
        y = bottom - (lat - min_lat) / (max_lat - min_lat) * (bottom - top)
        return x, y

    grid_font = ImageFont.truetype("arial.ttf", 17)
    note_font = ImageFont.truetype("arial.ttf", 18)
    for lon in range(-8, 5, 2):
        x, _ = xy(float(lon), min_lat)
        draw.line((x, top, x, bottom), fill="#E2E8F0", width=1)
        draw.text((x - 18, bottom + 12), f"{lon}°", font=grid_font, fill="#64748B")
    for lat in range(36, 45, 2):
        _, y = xy(min_lon, float(lat))
        draw.line((left, y, right, y), fill="#E2E8F0", width=1)
        draw.text((45, y - 10), f"{lat}°", font=grid_font, fill="#64748B")
    draw.rectangle((left, top, right, bottom), outline="#94A3B8", width=2)
    for _, row in valid.iterrows():
        x, y = xy(float(row["longitude_decimal"]), float(row["latitude_decimal"]))
        if bool(row["Y_Au"]) if pd.notna(row["Y_Au"]) else False:
            fill = "#D97706"
        elif bool(row["Y_Cu"]) if pd.notna(row["Y_Cu"]) else False:
            fill = "#0F766E"
        else:
            fill = "#475569"
        draw.ellipse((x - 3, y - 3, x + 3, y + 3), fill=fill)
    draw.text(
        (100, height - 45),
        f"n={len(valid):,} con coordenadas; CRS de salida: OGC:CRS84 (lon, lat). "
        "El rectángulo solo fija el marco del gráfico.",
        font=note_font,
        fill="#64748B",
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path, optimize=True)


def create_figures(clean: pd.DataFrame, tables: dict[str, pd.DataFrame], figures_dir: Path) -> None:
    figures_dir.mkdir(parents=True, exist_ok=True)
    commodities = tables["commodity_summary"].sort_values(
        "confirmed_present", ascending=False
    )
    draw_bar_chart(
        commodities["commodity_name"].tolist(),
        commodities["confirmed_present"].astype(int).tolist(),
        "Commodities confirmados en el subconjunto español",
        "Los estados desconocido y no informado no se cuentan como ausencia ni presencia",
        figures_dir / "commodity_counts.png",
        color="#B7791F",
        footnote=f"Base: {len(clean):,} registros seleccionados; un registro puede contener varios metales.",
    )
    techniques = tables["technique_summary"].sort_values(
        "confirmed_present", ascending=False
    )
    draw_bar_chart(
        techniques["technique_name"].tolist(),
        techniques["confirmed_present"].astype(int).tolist(),
        "Técnicas mineras confirmadas",
        "Recuentos positivos; los nulos se conservan como desconocidos",
        figures_dir / "technique_counts.png",
        color="#1B7F79",
        footnote=f"Base: {len(clean):,} registros seleccionados; categorías no mutuamente excluyentes.",
    )
    region = tables["region"].head(20)
    draw_bar_chart(
        region["region"].astype(str).tolist(),
        region["count"].astype(int).tolist(),
        "Concentración por región registrada en OxREP",
        "Top 20 categorías del campo region; no equivalen necesariamente a unidades administrativas actuales",
        figures_dir / "region_counts_top20.png",
        color="#386FA4",
    )
    accuracy = tables["coordinate_accuracy"].copy()
    accuracy["label"] = accuracy["coordinate_accuracy_reported"].map(
        lambda value: "No informado" if str(value) == "<Missing>" else str(value)
    )
    draw_bar_chart(
        accuracy["label"].tolist(),
        accuracy["count"].astype(int).tolist(),
        "Distribución de coordinateAccuracy",
        "Valor reportado por OxREP; el Excel no codifica unidad ni semántica de cero",
        figures_dir / "coordinate_accuracy.png",
        color="#7C3AED",
    )
    draw_map(clean, figures_dir / "spatial_distribution.png")


def source_missingness(source: pd.DataFrame) -> dict[str, int]:
    return {column: int(source[column].isna().sum()) for column in source.columns}


def write_audit_report(
    source: pd.DataFrame,
    clean: pd.DataFrame,
    decisions: pd.DataFrame,
    tables: dict[str, pd.DataFrame],
    duplicate_candidates_frame: pd.DataFrame,
    workbook_inventory: dict[str, Any],
    path: Path,
) -> None:
    direct_spain = int(
        source["country"].astype("string").str.strip().str.casefold().eq("spain").sum()
    )
    override_count = int(clean["flag_country_override"].sum())
    coords_complete = int(
        (clean["latitude_decimal"].notna() & clean["longitude_decimal"].notna()).sum()
    )
    commodity_table = tables["commodity_summary"]
    technique_table = tables["technique_summary"]
    top_combos = tables["commodity_combinations"].head(15)
    region_top = tables["region"].head(12)
    lines = [
        "# Auditoría reproducible de OxREP 3.0 y subconjunto España",
        "",
        "## Alcance y fuente",
        "",
        f"Se auditó programáticamente el libro `{workbook_inventory['workbook_file']}` "
        f"(SHA-256 `{workbook_inventory['workbook_sha256']}`), sin modificarlo. La hoja analítica "
        f"contiene {len(source):,} registros y {len(source.columns)} campos; la segunda hoja está vacía. "
        "OxREP indica que la versión 3.0 es una base de minas del mundo romano y advierte que la "
        "cobertura de hierro no es sistemática y que la datación es difícil para la mayoría de los sitios.",
        "",
        "Fuente institucional: https://oxrep.web.ox.ac.uk/mines-database (consulta: 2026-08-14).",
        "",
        "## Criterio exacto para España actual",
        "",
        f"1. Se normalizó únicamente espacio exterior y mayúsculas/minúsculas de `country`. "
        f"`country == Spain` produjo {direct_spain:,} registros.",
        f"2. Se añadió mediante override explícito `mineID=132` (Lanz), cuyo `country` crudo es "
        f"`Navarra`, coordenadas 42.99640694, -1.621914721 y hoja española 115. Se conserva "
        f"`country=Navarra` y `flag_country_override=true`.",
        f"3. Total curado: {len(clean):,} registros ({override_count} override). Todos los demás países "
        "quedan excluidos. El campo romano `province` no se usa para el filtro porque incluye sitios "
        "fuera de la España actual, especialmente Portugal.",
        "4. Las coordenadas se someten solo a un control amplio de plausibilidad. La validación "
        "administrativa point-in-polygon queda pendiente hasta incorporar un límite oficial reproducible.",
        "",
        "La tabla completa de inclusión/exclusión por registro está en "
        "`reports/audit/spain_selection_decisions.csv`.",
        "",
        "## Integridad del libro",
        "",
        f"- Hojas: {', '.join(workbook_inventory['sheet_order'])}.",
        "- La hoja principal es una tabla Excel A1:AU1400; no contiene fórmulas.",
        f"- `mineID`: {source['mineID'].notna().sum():,} informados, "
        f"{source['mineID'].nunique(dropna=True):,} únicos, rango "
        f"{int(source['mineID'].min())}-{int(source['mineID'].max())}. Los huecos de la secuencia "
        "son identificadores no presentes y no se imputan.",
        f"- Duplicados exactos de fila: {int(source.duplicated(keep=False).sum())}; "
        f"duplicados de `mineID`: {int(source['mineID'].duplicated(keep=False).sum())}.",
        "- No existe hoja de diccionario de datos ni campos de CRS/unidad para Lambert o "
        "`coordinateAccuracy`; la auditoría no inventa EPSG ni unidades.",
        "- Advertencia de lector: el render/importador de hojas puede mostrar `53` en celdas que "
        "en el XLSX son shared strings vacíos. El pipeline analítico usa pandas/openpyxl, que resuelven "
        "correctamente esas celdas; `53` no se incorpora a los datos.",
        "",
        "## Esquema y nulos",
        "",
        "Los 47 campos, tipos observados, nulos, cardinalidades y ejemplos están en "
        "`reports/audit/column_audit.csv`; todas las categorías y sus frecuencias, no solo el top-N, "
        "están en `reports/audit/value_counts_all_columns.csv`. Aspectos centrales:",
        "",
        f"- Coordenadas completas: {coords_complete:,}/{len(clean):,}; "
        f"faltan {len(clean)-coords_complete:,}.",
        f"- Geología informada: {clean['geology'].notna().sum():,}/{len(clean):,}; "
        f"tipo de depósito: {clean['depositType'].notna().sum():,}; explotación: "
        f"{clean['exploitationType'].notna().sum():,}.",
        f"- Alguna cronología: {int(clean['chronology_any_reported'].sum()):,}; "
        f"sin cronología: {int(clean['flag_chronology_missing'].sum()):,}.",
        "- Los indicadores mezclan texto (`TRUE/FALSE/?`), booleanos o 0/1 y nulos. El pipeline "
        "los normaliza a booleano nullable más un estado explícito: present, absent, unknown, missing o invalid.",
        "",
        "## Commodities",
        "",
        "| Código | Campo OxREP | Presente | Ausente | Desconocido | No informado |",
        "|---|---|---:|---:|---:|---:|",
    ]
    for row in commodity_table.itertuples(index=False):
        lines.append(
            f"| {row.commodity_code} | `{row.source_column}` | {row.confirmed_present} | "
            f"{row.confirmed_absent} | {row.unknown} | {row.missing} |"
        )
    lines.extend(
        [
            "",
            "Combinaciones principales (multi-label, no categorías exclusivas):",
            "",
            "| Combinación | Registros |",
            "|---|---:|",
        ]
    )
    for row in top_combos.itertuples(index=False):
        lines.append(f"| {row.commodity_combination} | {row.count} |")
    lines.extend(
        [
            "",
            f"Registros multimetálicos (>1 commodity confirmada): "
            f"{int(clean['confirmed_commodity_count'].gt(1).sum()):,}. La estructura real respalda "
            "una formulación multi-label. Las clases Fe, Sn, Hg y Zn son muy pequeñas para estimar "
            "modelos commodity-específicos robustos sin más etiquetas.",
            "",
            "## Técnicas",
            "",
            "| Técnica | Campo OxREP | Presente | Ausente | Desconocido/no informado |",
            "|---|---|---:|---:|---:|",
        ]
    )
    for row in technique_table.itertuples(index=False):
        lines.append(
            f"| {row.technique_name} | `{row.source_column}` | {row.confirmed_present} | "
            f"{row.confirmed_absent} | {row.unknown + row.missing + row.invalid} |"
        )
    lines.extend(
        [
            "",
            f"No tienen ninguna técnica confirmada {int(clean['flag_no_confirmed_technique'].sum()):,} "
            "registros; esto no significa ausencia de técnica porque la ficha puede estar incompleta.",
            "",
            "## Coordenadas y precisión",
            "",
            f"- Latitud observada: {clean['latitude_decimal'].min():.8f} a "
            f"{clean['latitude_decimal'].max():.8f}; longitud: {clean['longitude_decimal'].min():.8f} a "
            f"{clean['longitude_decimal'].max():.8f}.",
            f"- `coordinateAccuracy` no informado: "
            f"{int(clean['flag_coordinate_accuracy_missing'].sum()):,}; valor cero: "
            f"{int(clean['flag_coordinate_accuracy_zero'].sum()):,}; >1000: "
            f"{int(clean['flag_coordinate_accuracy_gt_1000'].sum()):,}.",
            "- Cero no se interpreta como precisión perfecta: aparece incluso junto a fuentes "
            "descritas como estimadas. La unidad y semántica no están codificadas en el libro.",
            "- La salida GeoJSON usa OGC:CRS84, orden longitud/latitud, porque los campos decimales "
            "son inequívocamente geográficos; no se transforma Lambert por falta de CRS documentado.",
            "",
            "## Depósitos, geología, explotación y cronología",
            "",
            "Las tablas completas están en `reports/tables/`. `geology` mezcla materiales, "
            "ambientes y narraciones; no se fuerza una taxonomía geológica. `depositType` contiene "
            "principal/secundario y valores inesperados; `exploitationType` contiene "
            "selectivo/extensivo y al menos un patrón compatible con desplazamiento de columnas. "
            "Se conservan los valores crudos y se añaden normalizaciones solo para vocabularios inequívocos.",
            "",
            "Los cuatro límites cronológicos son números de año, no fechas Excel. Los valores negativos "
            "se conservan sin reetiquetarlos como a. C. porque esa semántica no está formalizada en el libro. "
            "No se fabrican intervalos cuando faltan límites.",
            "",
            "## Distribución espacial",
            "",
            "| Región OxREP | Registros |",
            "|---|---:|",
        ]
    )
    for row in region_top.itertuples(index=False):
        lines.append(f"| {row.region} | {row.count} |")
    lines.extend(
        [
            "",
            "La fuerte concentración en el noroeste y varios distritos del sur combina señal histórica, "
            "geología, intensidad de explotación e historia de investigación. No debe interpretarse como "
            "muestra espacial aleatoria; condiciona background y validación.",
            "",
            "## Duplicados y casos problemáticos",
            "",
            f"- Duplicados de mineID en el subconjunto: "
            f"{int(clean['flag_duplicate_mine_id'].sum()):,} filas.",
            f"- Registros con coordenadas exactas compartidas: "
            f"{int(clean['flag_duplicate_coordinates'].sum()):,}; con nombre normalizado repetido: "
            f"{int(clean['flag_duplicate_site_name'].sum()):,}; con algún vecino <=100 m: "
            f"{int(clean['flag_nearby_record_100m'].sum()):,}.",
            f"- Se generaron {len(duplicate_candidates_frame):,} pares candidatos en "
            "`reports/audit/duplicate_candidates.csv`. Todos se retienen pendientes de revisión; cercanía "
            "o nombre repetido no demuestra identidad.",
            f"- Artefacto de texto detectado en {int(clean['flag_text_encoding_artifact'].sum()):,} "
            "registros; posible desplazamiento categorial en "
            f"{int(clean['flag_possible_column_shift'].sum()):,}.",
            "",
            "## Política de limpieza",
            "",
            "- No se modifica ni sobrescribe el libro original.",
            "- Se conserva cada campo OxREP crudo, `mineID`, fila Excel, hoja, hash del fichero y hash "
            "determinista de la fila.",
            "- No se elimina ningún registro problemático. Cada problema se expresa mediante flags "
            "booleanos documentados en `data/processed/quality_flag_definitions.csv`.",
            "- Nulo/desconocido no se convierte en cero. Esto es especialmente importante en "
            "commodities y técnicas.",
            "- Las categorías se normalizan solo en columnas derivadas; el valor original permanece.",
            "",
            "## Límites de esta fase",
            "",
            "OxREP documenta minas romanas y no es un inventario exhaustivo de mineralización. El Excel "
            "no permite inferir ley, tonelaje, rentabilidad, continuidad del cuerpo mineral, ausencia de "
            "mineralización ni etiquetas modernas de materias críticas. Tampoco permite validar relaciones "
            "geológicas causales sin integrar fuentes oficiales adicionales.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_data_dictionary(clean: pd.DataFrame, path: Path) -> None:
    rows = []
    raw_columns = {
        "source_excel_row",
        "source_sheet",
        "source_file",
        "source_file_sha256",
        "source_row_sha256",
    }
    for column in clean.columns:
        if column in raw_columns or column in {
            "mineID",
            "site",
            "ancientName",
            "province",
            "country",
            "region",
            "miningDistrict",
        }:
            role = "provenance_or_raw"
        elif column.startswith("flag_"):
            role = "quality_flag"
        elif column.startswith("Y_"):
            role = "provisional_target"
        elif column.startswith("has_technique_"):
            role = "normalised_indicator"
        elif column.endswith("_status"):
            role = "indicator_state"
        elif column.endswith("_clean") or column.endswith("_normalized"):
            role = "normalised_text"
        else:
            role = "raw_or_derived"
        description = ""
        if column in QUALITY_FLAG_DEFINITIONS:
            description = QUALITY_FLAG_DEFINITIONS[column][1]
        elif column == "Y_roman":
            description = "Positive-only indicator: all selected OxREP records are Roman-mine presences."
        elif column.startswith("Y_"):
            description = "Nullable multi-label commodity presence derived from the corresponding OxREP indicator."
        elif column == "coordinate_accuracy_reported":
            description = "Numeric copy of OxREP coordinateAccuracy; unit and zero semantics remain unasserted."
        rows.append(
            {
                "column": column,
                "dtype": str(clean[column].dtype),
                "role": role,
                "non_null": int(clean[column].notna().sum()),
                "unique_non_null": int(clean[column].nunique(dropna=True)),
                "description": description,
            }
        )
    write_csv(pd.DataFrame(rows), path)


def output_manifest(root: Path, paths: list[Path], pipeline_version: str) -> dict[str, Any]:
    records = []
    for path in sorted(set(paths)):
        if path.exists() and path.is_file():
            records.append(
                {
                    "relative_path": path.relative_to(root).as_posix(),
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
    return {
        "pipeline_version": pipeline_version,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "python": sys.version,
        "platform": platform.platform(),
        "pandas": pd.__version__,
        "numpy": np.__version__,
        "openpyxl": openpyxl.__version__,
        "pillow": getattr(Image, "__version__", "unknown"),
        "files": records,
    }


def refresh_output_manifest(root: Path, pipeline_version: str = "0.2.0") -> int:
    audit_dir = root / "reports" / "audit"
    output_bases = [
        audit_dir,
        root / "reports" / "tables",
        root / "reports" / "figures",
        root / "data" / "processed",
        root / "data" / "interim",
        root / "docs",
        root / "config",
        root / "src",
        root / "tests",
        root / "tools",
        root / "notebooks",
    ]
    tracked_outputs = [
        path
        for base in output_bases
        for path in base.rglob("*")
        if path.is_file()
        and path.name != "output_manifest.json"
        and "__pycache__" not in path.parts
        and path.suffix != ".pyc"
    ]
    tracked_outputs.extend(
        path
        for path in [
            root / ".gitignore",
            root / "README.md",
            root / "pyproject.toml",
            root / "requirements.txt",
            root / "run_pipeline.py",
        ]
        if path.is_file()
    )
    tracked_outputs.extend(sorted((root / "reports").glob("*.pdf")))
    write_json(
        audit_dir / "output_manifest.json",
        output_manifest(root, tracked_outputs, pipeline_version),
    )
    return len(tracked_outputs) + 1


def run_pipeline(root: Path) -> None:
    config = json.loads((root / "config" / "project.json").read_text(encoding="utf-8"))
    workbook_path = root / config["input_workbook"]
    report_path = root / config["input_report"]
    methodological_references: list[dict[str, Any]] = []
    for configured_reference in config.get("methodological_references", []):
        reference_path = root / configured_reference["path"]
        if not reference_path.is_file():
            raise FileNotFoundError(f"Methodological reference not found: {reference_path}")
        observed_hash = sha256_file(reference_path)
        expected_hash = configured_reference.get("expected_sha256")
        if expected_hash and observed_hash.casefold() != expected_hash.casefold():
            raise ValueError(
                f"Methodological reference hash changed for {reference_path.name}: "
                f"{observed_hash} != {expected_hash}"
            )
        methodological_references.append(
            {
                "path": reference_path,
                "role": configured_reference["role"],
                "expected_sha256": expected_hash,
            }
        )
    audit_dir = root / "reports" / "audit"
    tables_dir = root / "reports" / "tables"
    figures_dir = root / "reports" / "figures"
    processed_dir = root / "data" / "processed"
    for directory in [audit_dir, tables_dir, figures_dir, processed_dir]:
        directory.mkdir(parents=True, exist_ok=True)

    source, workbook_inventory = inspect_workbook(
        workbook_path,
        report_path,
        config["source_sheet"],
        audit_dir,
        methodological_references,
    )
    expected_shape = (config["expected_source_rows"], config["expected_source_columns"])
    if source.shape != expected_shape:
        raise ValueError(f"Source shape {source.shape} != configured {expected_shape}")
    (
        clean,
        decisions,
        commodity_long,
        technique_long,
        duplicate_candidates_frame,
        quality_definitions,
    ) = prepare_spain_dataset(
        source, config, workbook_inventory["workbook_sha256"]
    )

    write_csv(decisions, audit_dir / "spain_selection_decisions.csv")
    write_csv(duplicate_candidates_frame, audit_dir / "duplicate_candidates.csv")
    write_csv(quality_definitions, processed_dir / "quality_flag_definitions.csv")
    write_csv(clean, processed_dir / "oxrep_roman_mines_spain_clean.csv")
    write_csv(commodity_long, processed_dir / "oxrep_spain_mine_commodities.csv")
    write_csv(technique_long, processed_dir / "oxrep_spain_mine_techniques.csv")
    write_data_dictionary(clean, processed_dir / "data_dictionary.csv")
    export_geojson(clean, processed_dir / "oxrep_roman_mines_spain_clean.geojson")
    export_sqlite(
        clean,
        commodity_long,
        technique_long,
        quality_definitions,
        duplicate_candidates_frame,
        processed_dir / "oxrep_roman_mines_spain.sqlite",
    )
    tables = build_summary_tables(clean, commodity_long, technique_long, tables_dir)
    create_figures(clean, tables, figures_dir)
    write_audit_report(
        source,
        clean,
        decisions,
        tables,
        duplicate_candidates_frame,
        workbook_inventory,
        root / "docs" / "oxrep_audit.md",
    )

    validation = {
        "checks": {
            "source_shape_matches": list(source.shape) == list(expected_shape),
            "selected_rows": int(len(clean)),
            "mineID_complete": bool(clean["mineID"].notna().all()),
            "mineID_unique": bool(clean["mineID"].is_unique),
            "source_rows_unique": bool(clean["source_excel_row"].is_unique),
            "selection_decision_rows_equal_source": len(decisions) == len(source),
            "selection_true_equals_output": int(decisions["include_spain_subset"].sum())
            == len(clean),
            "Y_roman_all_one": bool(clean["Y_roman"].eq(1).all()),
            "complete_coordinate_records": int(
                (
                    clean["latitude_decimal"].notna()
                    & clean["longitude_decimal"].notna()
                ).sum()
            ),
            "geojson_feature_count": len(
                json.loads(
                    (
                        processed_dir / "oxrep_roman_mines_spain_clean.geojson"
                    ).read_text(encoding="utf-8")
                )["features"]
            ),
            "geojson_count_reconciles": len(
                json.loads(
                    (
                        processed_dir / "oxrep_roman_mines_spain_clean.geojson"
                    ).read_text(encoding="utf-8")
                )["features"]
            )
            == int(
                (
                    clean["latitude_decimal"].notna()
                    & clean["longitude_decimal"].notna()
                ).sum()
            ),
            "commodity_long_rows": int(len(commodity_long)),
            "technique_long_rows": int(len(technique_long)),
            "quality_flag_columns_present": all(
                column in clean.columns for column in QUALITY_FLAG_DEFINITIONS
            ),
            "input_workbook_hash_unchanged": sha256_file(workbook_path)
            == workbook_inventory["workbook_sha256"],
            "methodological_reference_hashes_match_config": all(
                item["expected_sha256"] is None
                or sha256_file(item["path"]).casefold()
                == item["expected_sha256"].casefold()
                for item in methodological_references
            ),
        }
    }
    write_json(audit_dir / "validation_report.json", validation)

    output_count = refresh_output_manifest(root)
    print(
        json.dumps(
            {
                "source_shape": source.shape,
                "selected_records": len(clean),
                "complete_coordinates": int(
                    (
                        clean["latitude_decimal"].notna()
                        & clean["longitude_decimal"].notna()
                    ).sum()
                ),
                "outputs": output_count,
            },
            indent=2,
        )
    )
